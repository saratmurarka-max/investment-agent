import asyncio
import base64
import io
import re
from collections import defaultdict
from datetime import date as date_type, datetime

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from backend.db.database import get_db
from backend.db.models import Client, Holding, Portfolio, RealizedPnL, DerivativeTrade
from backend.services import market_data

router = APIRouter(prefix="/portfolios", tags=["portfolios"])

# PROFITMART logo embedded as base64 PNG (source: profitmart.in)
_PROFITMART_LOGO_B64 = (
    "iVBORw0KGgoAAAANSUhEUgAAAS8AAAA8CAYAAAA36zxZAAAQsElEQVR42uydCXQURRqAK5JE5EYW"
    "ERQSBTQGhBwkmZBjsoAIiKBiCFGOuOQAMcH1Am/FXTzYfSFRlxBcXA6v1Sce+HTBNRAggAgeiwos"
    "rLjCIiirgrBAEmb/mvl7ptNH9VUzScbive9NZqbq75qe7m/+qq5qyMLN2cEmFXgZiOIQyxJVu9LI"
    "Q0vHkam5s0hRYREpKhIIBOGCY0EYcAXwE+ABtgKRHGIKeQkEgqDKqx/wM+CRsR1owyG2kJdA8AvH"
    "sSAY4voR8GiwJVQZmJCXQBC+OBaEBpdJXUUGW0ORgQl5CQThi2NBKOjv7yoasy3YGZiQl0AQvjgW"
    "hFHG1ZwZmJCXQBC+OBaEMuOywTYO2xfyEgh+YTgWBBAHHAM8DvgwGF1IIS+BIHwhHDIu71VFDny0"
    "kIOwhLwEgl8GzZ1xaWVgEQs5iEvISyAIb2zPnOeYcfnhPRNfyEsgCF9sTUANQsalZDuPDEzISyAI"
    "X1pSxsV9GoWQl0AQvlgSl82MazMwETjJKheMpURCXgJB+GJ5kbVFvgTaY4wkoN5MPV6LuRfvSSX3"
    "V11HJl9XSoqLC+3soLbAJUAsEKNDLNIDiLIY/3zgUpOxo4s4fOE6bcgCJgK3ANOACYAL6GAz5nkm"
    "9luMic8dxYwZ+DtSVu4CaZ9ygsa6kMN+FnAmmF3FL4EOiljJwOlQZWB/2plOnnhzJJk6eSaZPH0m"
    "mVpS4mc6fPhi4x10FeAxyWngELABqAJuANoYxK+wEPsw8CGwvMj3r5eDL747cDvwHnCcsd2jwCqg"
    "AGhrIf4YWp8D8SZj9paVe5fDdpVs0viMGcDoIg4noQ4DgGuA8x3EGIQxOjtpB+77a4FxMsbja8Nx"
    "O10N4vTH8mM1Cbw32rs947IjgpVx7fOLS01yKDOwxbvTyF0L8kj+xFJyS3GxD5DXTTNmmJHXaIcH"
    "/b+AKYz4ix3EPg08CURYPBgfBn62sb2DQLHJbYzjIAzKQJMxY2TlajlsV8knGp9xN773OAdRKWkn"
    "2/YImzHaAGcxxrMO2vKyhf20lXG8P8fhe2gCU1wOMq6uBlJJAxqDnoHVZZN5X2eR1VVZ5GB8DjmU"
    "7SaH3EB6OqmcNImMLi0lJYWFvDIvFkv0Mi8OsT8x+et8BfAxh+29DXRqQZnX2WbKvGpk73fjICw5"
    "z8tiZ9uMkS+LcdxyfbV05gOj8HsYI8uQJgFz/Ps9sL8u0MgCCxhMAeZh/X3AVGAao3wuz0XWlD1A"
    "B7k8yrdkkwU7NLeRApwK9lXI+/bnkNryocTTO5F4kpKJJxkYOJAczc4m1996K+1ChkJelGeDJC/K"
    "FwYHYBpwhsN2JL6iB2cI5JVgQ17bOGxXyR4pvk6G9zcOwpL/yHg4yGs3cAJ4EuNMthmnGuu7jMqi"
    "xJdh+f3AuRa39Susu8ZMeZ6LrL8A2svFVQE8/g83+QMwf6ebVG5UbStRmYHxvp3Og1+5ydqFmcTT"
    "N4140lzE40IGDyaLJ04kV5WVkRIr8gqcMK/jF7UcWAGsBY4YnARJJuX1b1nslfhlfmMQe5bOZ+hF"
    "28usGzjYtmN2doBZNvDrGGFRXg1AHZ74G0wQa0Nez+Nn2afEYB9+o1UHY72qI4c9sswknYO4KJ9K"
    "2To+um3EiMe61fS51O13KK8JFuq8gnWesritVKy30Y68+vLKuCqAuYfcZNWKLHIkw0WWvZNF7j2g"
    "KbAUoDFYGZimvKDb6ElMJN9BF3KCL/uyKq9GxljFAsYJstqkvFbqjGHcxoi9TadNOxl1TgJPAAOU"
    "28IDaZGBwF6xKK9jyrIcLgIE5GVMvNH4mgX24gWaSEl+HMQ1EmPdAmQ6yLxewLqp+HwVPh8UInl1"
    "wTon6LEUCnnZXWS9u2nG5WMOiOvFVVnEEw/S6JhKTqW4SNX6bJbAGoJxOx21vJpmX8/l5vqyr8JC"
    "q/LqaXzwBJB9mVEm5PU6I/Zc3Tapx75KGSfrbmmg2wAXcJIRJ8uCaI4D5zWjvGIZn+NSi+3ZQ+vh"
    "3wX492yH8vpRFjPfprwisN53stcS8LW3QiQvyn5pvwZbXnYXWX8OdFZlXJK4+oMw+gFuEEVsGmlI"
    "cJHFNSCwg5oCSwAaeN9OhymvxETv2NeNvuzLqrwuZuzYyxknyWAT8lrFiN0BaNCpN0RWLhI4wciA"
    "Olo4qAYzPs9mi/Lq3IzyGsj4HIPsygv5Hp9H2fx8s7H+rfh8jk15zcB6dypeP4qvtw2RvGqloZJg"
    "ymsY8JPNjKudVsa18i0Q1+UorgyvKHyPfdJIfRJkYLXMDMxjgx2+zLFpvIpNPlG+9WeQ16UqeanH"
    "vgoLecmrDWP+1EiH8opmTHVIlpW7XquMg8HbBawul5CX/1hZamdag39eXeC1u2zJKzCu11knCy8N"
    "kbw+wXpxwZTXu3YG51UZl7+rqBSXH9/zGBTYeioWrl3IO5qKy3eV8+4jOeTjezK88gq0penY1/eQ"
    "feX6si9e8moHnNKp+2uH8uri3b66Tr3igF2hE/uITXl0Ypz4Dwh5edmEr/W1GGupxpyuO23IK40x"
    "7HAuvvdDiOR1GutFB1NeXYAPLM7j6qQrrstAEn1V4lILDLqQ1TVZKDDHE1krgTZKcc351k22PJLh"
    "7bJ6UqQ2aGdfi/LyyAhf9sVDXm7GSdLHobxuV5VXf+HnAId0ypVbFofxBND3TIrmp2YesA+2vHrg"
    "a9stxOkpnwjrUF6rmwzUq3lRGl4IsrymYZ3nQ3W1cYPJjKsDU1z9NMUV1eR5pq8L2ZDAHMRPNnkV"
    "soKWV4srh2x6DMR1IbRnMGxvqKpNEfKxryOBK4+m5GWw3q1Gp95hk1cbl+vEHcaY9jBNVu5i+hqr"
    "nE0q9Qb/TYrmFFAGTAeKGcS0UnnJT/axJuNslLI1h/JqpzrG1AzAMu8EUV79Zd9L11DJi/KBo4yr"
    "H5CpkkQ+8A0wTCGwQBeyVrcL6QLqzYqrEsT1FGZcm+ehuBI0xbUa2A1EyLOvJYGxL0N56VwCjpMm"
    "6emwyKS8PsKrTTcBU3Aw9w1G3J3K2cyMsuMcyOsRnZj/Ac7hOEn1plYsrwjaJpPTQoZgjJfoc4fy"
    "msuY7xcgkJF3CsIk1TLZ/sygr4VSXpQaLXHZzLgmAB6kEcjR7EIm0i4kzIbXz8BOa3UVlRnXUzt8"
    "Vzrr5jEzriWyNm3xZoXp6f4rj3kzvQu4zUxSrQXWIGuBHSZOyh5BmGF/QGO2ezKj/CgH8rqXMY4W"
    "zVFe41uxvCgl+N4jBjG+lq4AcpDXf7F8e4Ny0zXbZrz+9o/AjXj3kTzgOuAGHMZYK1/fKA3SN4e8"
    "KBuU4rKRcY0HPBpkagms0aALqcjAyvUyrjraVeyhm3Et0WjPDiBCyr6qYOxrOGRfMwoLeS8PmhmE"
    "5UF/xztEEAuZ1/hWkHmNa+XyUl/1U/MbuUQcymukfHKzAZGBhf3Gn9NwonJggvRLHFYZcJEXZRdw"
    "1GbGdS3g0eEskKV9FdJQYKeAl1RjXPKMq6d5cSkysAhp7CvXl33xlNf9nBdmHwdyjQaBdShwcHA9"
    "oxNzF+e1jVPCQF4p6lUVAbDdp+jfHOS1Dssmmmz7X7F8joXM626c75eEjyNk0yGyOKwu4Cqv84Bu"
    "SnHdA9nNC2/QCagGGZcxLn2B6XYhezfJuKTB+cPmMy5DgQ0a5J33NXz2bJp9OZXX+0Cmjft51QPH"
    "GDPbzxhdyWGs43vGwcFVpxPzHaVoDG7j8ygwj0FcGMiL8r5O7Cfx9Zs5yKs7lltnsf2NJiVRzWhH"
    "e9lE6LiWJC85XpE88DVMQH0TxBXnnQ5hnHEZk2VDYIGu4se+ruLG3zMzrioL7dkC2VdbetucfMi+"
    "ppWUsOT1LXAAu0wHcTnEZ/hL+6DJAcsKxhrIjrhKfyqjDesNb6ui5qjNA6srox33iqkSPnSmTvxT"
    "Yz/uo885yOt3sjuLLANeNEJxH7leDq82JsmOq8gWKa/HdrnJilWSuDQzruGAxwZDrQqs0vzg/GLL"
    "7UlP3wvZV9sXJkwgV5eWsq42xuIVx0gHX1qFyQXP5TZmy1/DqFNoo61PM+LFiUmqujyhuIL6Hj4f"
    "zEleR2g56dECB/FxPoepEiVS9tci5UXHobY+BPfF6p6KklDRD/goGALzjYEp53GhuHqk6mVcz9lq"
    "S2rqw54hQyLnT55M7/fFkldHDmmy2Umq5/hnLKs5wbhv0g+Mbmd3C+3MYJwE68TaRkMaMEPPlA0n"
    "EA7yGmV5MmiA9tLxw0Fe8psWLmxx8npgv5vULABZXCKtEdRlU7AE9nStNMaVA11FaR4Xp4zLl3XN"
    "pGNedWPGkBEw5lVsMMM+hPLy/7rp8Be9y+IGt6iON9HGYUA9I45LyMuQm7FsAz724CQvaXrOFTb3"
    "6XLpCjQHeVE+x7I3tlZ5UdZzF1gNXafopt1InDnPZYxLznRPSgqBrIvMLSgg42fNamny8p8wzEXZ"
    "aj406D5U4kHTVTEInAOsMKi7jJYX8jLFDnk3jYO8YvH9Lx0cf5dhjE85yauL7Ieub2uVF1+B0dvp"
    "pLi8t5ReU5HJGuNaaldcdKIqzbo2jxlDZ9n7dlDLk1ca4+TbxTigjpkYA/kZ2Iu3eP6fifKf0fhC"
    "XubKYxtL6N+c5LUE388zG9PgZpWXc5CXfIjhMBDZWuVFqeMmMHqhIBUek4BErhnXTG/91FRv1nUP"
    "ZF3jaNbVMuVFWcs4AcsYJ+0xDvPVJPYCXYS8ijwcjgG78mqk73O4wicton6Wk7wod0hje61ZXvwy"
    "sKHAICCJ4xgXZFy0vpR1bRs1iowsK+NxV4lgyos1AbUe6MT45a/jIK7XcBEwEfJqNnnl4XtLOWwn"
    "GjiDRHOSF+VVqZvcmuXlRGCZ6licriq6XIVSDGms62461nXbbaTY3P/b2IfDgVOtE3uNQb1yy3O/"
    "AvxWmsBqkZ0WbmI4Xj+OOmPjEDOGwx1hEy2259sgy+s+jD9M472j+N4lHLYjXwJ0u8Z7K/G9SXb+"
    "ByOsO9Vk+XRpDK4lyYuy0YZgzkoZmA7VNsU1o0mchARSM3YsubqsTPkf0Y7AK0T1Ck5zyrzKlfHx"
    "+WvMeoGD6oSyvsnbsZyLa+velu6brsP3+J+OTrT4uUbrfK4f5JkXh5iUMxa+iwH+emqutNieL2g9"
    "DscA60emEcjSWXa0hsM25LdQOqkzbroIs9sbbP3P7IH5ZwNMlE/Bz7y+pckrgnMGVuU446KkwedJ"
    "SiIL8/PppFRSrD7JLwJ6yuiFj204HDSdNOJfZOE+SN006scAfSy24Ur8hb8e7xbgxmkU7UzHMbff"
    "LgQiOMe08l1EYp1e8hhIlI0TsyeHY0CP9iiVaI3t9uYQX/V5dKZcdMbttXXwvcV797Fx2Sj8zN1a"
    "mrwk1tnMwHKUaxVtUOytr5aXdz3jNdBl5HAQCASCENAc8qJssCmfDOBhm3ULAttXy6tayEsgaFU0"
    "l7z8AgsR3oxLyEsgCB+aU16hEph3OoSQl0AQXjS3vIItMO8EVCEvgSD8aAnyomziICrNjEvISyAI"
    "T1qKvCjrOQiribiEvAT/b9+OUQCEgSCKelgLPYVHE0FyLrOCrRZuwkz4xWuT7jMpgnEpxStrga3V"
    "RLyAsanFK5Qui4t4AdYU4xWOLuEiXoAt1XiF0uypSLwAe8rxCmezxUW8AGvq8fr6zD3/voN4AZbU"
    "4/W2wJaEc4kXYMolXvcCS11cxAuw5hKvx15tCecQL8DcBfWvv2p2AQg3AAAAAElFTkSuQmCC"
)
_PROFITMART_LOGO_BYTES: bytes = base64.b64decode(_PROFITMART_LOGO_B64)


# --- Schemas ---

class HoldingIn(BaseModel):
    ticker: str
    shares: float
    avg_cost: float


class PortfolioIn(BaseModel):
    name: str
    currency: str = "USD"
    holdings: list[HoldingIn] = []


class ClientIn(BaseModel):
    name: str
    email: str
    risk_tolerance: str = "medium"


# --- Helpers ---

def _normalise_headers(row) -> list[str]:
    return [str(h).lower().strip().replace(" ", "_").replace("\\", "") if h else "" for h in row]


def _to_exchange_ticker(symbol: str) -> str:
    """
    Resolve a raw broker symbol to a yfinance-compatible ticker.
    Defaults to .NS (NSE) — the primary exchange for most Indian stocks.
    Symbols already carrying .NS / .BO are returned unchanged.
    Numeric BSE codes (e.g. 531637-EQ) are converted to BSE format (531637.BO).
    """
    symbol = symbol.strip().upper()
    # Already qualified
    if symbol.endswith(".NS") or symbol.endswith(".BO"):
        return symbol
    # Numeric BSE code (e.g. 531637-EQ or 531637) → use BSE format for yfinance
    if re.match(r"^\d+(-EQ)?$", symbol):
        base = re.sub(r"-EQ$", "", symbol)
        return base + ".BO"
    # Default to NSE (primary Indian exchange)
    return symbol + ".NS"


# ── Equity broker format ────────────────────────────────────────────────────────

def _is_broker_format(rows: list) -> bool:
    if len(rows) < 4:
        return False
    headers = _normalise_headers(rows[3])
    # Derivative files also have scrip_symbol + purchase_qty — exclude them
    if "instrument_type" in headers:
        return False
    has_symbol = "scrip_symbol" in headers or "scrip_name" in headers
    return has_symbol and "purchase_qty" in headers


def _extract_client_info(rows: list, headers: list[str]) -> dict:
    """
    Extract broker client_id and client_name from the first data row.
    Both equity and derivative PROFITMART files have Client_id and Client name columns.
    Returns {"client_id": "17010040", "client_name": "SWATI SURESH GAIKWAD"} or {}.
    """
    cid_idx  = headers.index("client_id")   if "client_id"   in headers else -1
    cname_idx = headers.index("client_name") if "client_name" in headers else -1

    info: dict = {}
    for row in rows[4:]:  # first data row is enough
        try:
            if cid_idx >= 0 and row[cid_idx] is not None:
                raw = str(row[cid_idx]).strip()
                # strip decimal suffix if it came through as float (e.g. "17010040.0")
                if raw.endswith(".0"):
                    raw = raw[:-2]
                if raw and raw.upper() not in ("NONE", ""):
                    info["client_id"] = raw
            if cname_idx >= 0 and row[cname_idx] is not None:
                raw = str(row[cname_idx]).strip()
                if raw and raw.upper() not in ("NONE", ""):
                    info["client_name"] = raw
        except Exception:
            pass
        if info:
            break  # found on first valid row — stop
    return info


def _parse_broker_format(rows: list) -> tuple[list[dict], list[dict], list[str], dict]:
    """
    Returns:
        holdings      - list of open positions {ticker, shares, avg_cost}
        realized_pnls - list of realized P&L per ticker {ticker, short_term_gain, long_term_gain}
        skipped       - list of skipped row descriptions
        client_info   - {"client_id": ..., "client_name": ...} extracted from broker file
    """
    if len(rows) < 5:
        return [], [], [], {}

    headers = _normalise_headers(rows[3])

    def col(names: list[str]) -> int:
        for n in names:
            if n in headers:
                return headers.index(n)
        return -1

    sym_idx       = col(["scrip_symbol"])          # NSE/BSE ticker symbol (e.g. RELIANCE, BANKBARODA)
    name_idx      = col(["scrip_name"])            # Company display name (e.g. Reliance Industries Ltd.)
    qty_idx       = col(["purchase_qty"])
    rate_idx      = col(["purchase_rate"])
    sell_qty_idx  = col(["sell_qty"])
    stcg_idx      = col(["shorterm_pl", "short_term_pl", "shorterm_p\\l"])
    ltcg_idx      = col(["actual_longterm", "longterm_pl", "actual_longterm_pl"])

    # If no scrip_symbol column, fall back to scrip_name as ticker
    if sym_idx == -1:
        sym_idx = name_idx

    if any(i == -1 for i in [sym_idx, qty_idx, rate_idx]):
        return [], [], ["Could not find required broker columns"], {}

    client_info = _extract_client_info(rows, headers)

    open_agg: dict[str, dict] = defaultdict(lambda: {"net_qty": 0.0, "cost_basis": 0.0, "name": ""})
    real_agg: dict[str, dict] = defaultdict(lambda: {"short_term": 0.0, "long_term": 0.0})
    skipped: list[str] = []

    for i, row in enumerate(rows[4:], start=5):
        try:
            symbol = str(row[sym_idx] or "").strip().upper()
            if not symbol or symbol == "NONE":
                continue

            display_name = str(row[name_idx] or "").strip() if name_idx >= 0 else ""

            buy_qty  = float(row[qty_idx]  or 0)
            buy_rate = float(row[rate_idx] or 0)
            sell_qty = float(row[sell_qty_idx] or 0) if sell_qty_idx >= 0 else 0

            if buy_qty <= 0:
                continue

            # Realized P&L for sold portion
            if sell_qty > 0:
                stcg = float(row[stcg_idx] or 0) if stcg_idx >= 0 else 0
                ltcg = float(row[ltcg_idx] or 0) if ltcg_idx >= 0 else 0
                real_agg[symbol]["short_term"] += stcg
                real_agg[symbol]["long_term"]  += ltcg

            # Open position (unsold portion)
            net_qty = buy_qty - sell_qty
            if net_qty > 0:
                open_agg[symbol]["net_qty"]    += net_qty
                open_agg[symbol]["cost_basis"] += net_qty * buy_rate
                if display_name and not open_agg[symbol]["name"]:
                    open_agg[symbol]["name"] = display_name

        except Exception:
            skipped.append(f"row {i}")

    holdings = [
        {
            "ticker":   _to_exchange_ticker(sym),
            "name":     data["name"] or None,
            "shares":   round(data["net_qty"], 6),
            "avg_cost": round(data["cost_basis"] / data["net_qty"], 4),
        }
        for sym, data in open_agg.items()
        if data["net_qty"] > 0
    ]

    realized_pnls = [
        {
            "ticker":           _to_exchange_ticker(sym),
            "short_term_gain":  round(data["short_term"], 4),
            "long_term_gain":   round(data["long_term"], 4),
        }
        for sym, data in real_agg.items()
        if data["short_term"] != 0 or data["long_term"] != 0
    ]

    return holdings, realized_pnls, skipped, client_info


# ── Derivative format ───────────────────────────────────────────────────────────

# Matches: "IO CE NIFTY 02Dec2025 26600" or "IO PE NIFTY 31Jul2025 24300"
_DERIV_RE = re.compile(
    r"^(?:IO|ST|IX)\s+(CE|PE|FU|CA|PA)\s+(\w+)\s+(\d{2}\w{3}\d{4})(?:\s+(\d+(?:\.\d+)?))?",
    re.IGNORECASE,
)


def _parse_scrip_symbol(symbol: str) -> dict:
    """Extract option_type, underlying, expiry_date, strike_price from a scrip symbol."""
    m = _DERIV_RE.search(symbol.strip())
    if not m:
        return {"option_type": None, "underlying": None, "expiry_date": None, "strike_price": None}
    option_type = m.group(1).upper()
    underlying  = m.group(2).upper()
    expiry_str  = m.group(3)
    strike_str  = m.group(4)
    try:
        expiry = datetime.strptime(expiry_str, "%d%b%Y")
    except Exception:
        expiry = None
    strike = float(strike_str) if strike_str else None
    return {
        "option_type":  option_type,
        "underlying":   underlying,
        "expiry_date":  expiry,
        "strike_price": strike,
    }


def _is_derivative_format(rows: list) -> bool:
    """Detect PROFITMART-style Derivative P&L Excel (DER P&L report)."""
    if len(rows) < 4:
        return False
    headers = _normalise_headers(rows[3])
    return "instrument_type" in headers and any(
        h in headers for h in ["booked_p/l", "booked_pl", "booked_profit"]
    )


def _parse_derivative_format(rows: list) -> tuple[list[dict], list[str], dict]:
    """
    Parse a PROFITMART derivative P&L Excel.
    Returns:
        trades      - list of trade dicts ready to be inserted as DerivativeTrade rows
        skipped     - list of skipped row descriptions
        client_info - {"client_id": ..., "client_name": ...} extracted from the file
    """
    if len(rows) < 5:
        return [], [], {}

    headers = _normalise_headers(rows[3])

    def col(names: list[str]) -> int:
        for n in names:
            if n in headers:
                return headers.index(n)
        return -1

    sym_idx    = col(["scrip_symbol"])
    inst_idx   = col(["instrument_type"])
    tdate_idx  = col(["trade_date"])
    bqty_idx   = col(["purchase_qty"])
    brate_idx  = col(["purchase_rate"])
    bamt_idx   = col(["purchase_amount"])
    sdate_idx  = col(["sell_trade_date"])
    sqty_idx   = col(["sell_qty"])
    srate_idx  = col(["sell_rate"])
    samt_idx   = col(["sell_amount"])
    pnl_idx    = col(["booked_p/l", "booked_pl"])
    profit_idx = col(["booked_profit"])
    loss_idx   = col(["booked_loss"])

    if sym_idx == -1:
        return [], ["Could not find Scrip_Symbol column"], {}

    client_info = _extract_client_info(rows, headers)

    def to_float(idx: int, row) -> float:
        if idx < 0 or idx >= len(row):
            return 0.0
        v = row[idx]
        if v is None:
            return 0.0
        try:
            return float(v)
        except Exception:
            return 0.0

    def to_dt(idx: int, row):
        if idx < 0 or idx >= len(row):
            return None
        v = row[idx]
        if v is None:
            return None
        if isinstance(v, datetime):
            return v
        try:
            return datetime.strptime(str(v), "%Y-%m-%d")
        except Exception:
            return None

    trades: list[dict] = []
    skipped: list[str] = []

    for i, row in enumerate(rows[4:], start=5):
        try:
            symbol = str(row[sym_idx] or "").strip()
            if not symbol or symbol.upper() == "NONE":
                continue

            parsed = _parse_scrip_symbol(symbol)
            inst   = str(row[inst_idx] or "").strip() if inst_idx >= 0 else None

            trades.append({
                "scrip_symbol":    symbol,
                "instrument_type": inst,
                "option_type":     parsed["option_type"],
                "underlying":      parsed["underlying"],
                "expiry_date":     parsed["expiry_date"],
                "strike_price":    parsed["strike_price"],
                "trade_date":      to_dt(tdate_idx, row),
                "buy_qty":         to_float(bqty_idx, row),
                "buy_rate":        to_float(brate_idx, row),
                "buy_amount":      to_float(bamt_idx, row),
                "sell_date":       to_dt(sdate_idx, row),
                "sell_qty":        to_float(sqty_idx, row),
                "sell_rate":       to_float(srate_idx, row),
                "sell_amount":     to_float(samt_idx, row),
                "booked_pnl":      to_float(pnl_idx, row),
                "booked_profit":   to_float(profit_idx, row),
                "booked_loss":     to_float(loss_idx, row),
            })
        except Exception:
            skipped.append(f"row {i}")

    return trades, skipped, client_info


# --- Routes ---

@router.post("/clients")
async def create_client(body: ClientIn, db: AsyncSession = Depends(get_db)):
    client = Client(**body.model_dump())
    db.add(client)
    await db.commit()
    await db.refresh(client)
    return {"id": client.id, "name": client.name, "email": client.email}


@router.get("/clients/{client_id}/portfolios")
async def list_portfolios(client_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Portfolio)
        .options(selectinload(Portfolio.holdings))
        .where(Portfolio.client_id == client_id)
    )
    portfolios = result.scalars().all()
    return [
        {
            "id": p.id,
            "name": p.name,
            "currency": p.currency,
            "holding_count": len(p.holdings),
        }
        for p in portfolios
    ]


@router.post("/clients/{client_id}/portfolios")
async def create_portfolio(
    client_id: int, body: PortfolioIn, db: AsyncSession = Depends(get_db)
):
    portfolio = Portfolio(client_id=client_id, name=body.name, currency=body.currency)
    db.add(portfolio)
    await db.flush()
    for h in body.holdings:
        db.add(Holding(portfolio_id=portfolio.id, **h.model_dump()))
    await db.commit()
    await db.refresh(portfolio)
    return {"id": portfolio.id, "name": portfolio.name}


@router.get("/{portfolio_id}")
async def get_portfolio(portfolio_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Portfolio)
        .options(selectinload(Portfolio.holdings))
        .where(Portfolio.id == portfolio_id)
    )
    portfolio = result.scalar_one_or_none()
    if not portfolio:
        raise HTTPException(404, "Portfolio not found")

    return {
        "id": portfolio.id,
        "name": portfolio.name,
        "currency": portfolio.currency,
        "holdings": [
            {
                "id": h.id,
                "ticker": h.ticker,
                "name": h.name,
                "shares": float(h.shares),
                "avg_cost": float(h.avg_cost),
            }
            for h in portfolio.holdings
        ],
    }


@router.get("/{portfolio_id}/pnl")
async def get_portfolio_pnl(portfolio_id: int, db: AsyncSession = Depends(get_db)):
    """
    Returns per-holding P&L (unrealized) using live prices + total realized P&L.
    """
    result = await db.execute(
        select(Portfolio)
        .options(selectinload(Portfolio.holdings))
        .where(Portfolio.id == portfolio_id)
    )
    portfolio = result.scalar_one_or_none()
    if not portfolio:
        raise HTTPException(404, "Portfolio not found")

    holdings = portfolio.holdings
    if not holdings:
        return {
            "holdings": [],
            "summary": {
                "total_invested": 0,
                "total_current_value": 0,
                "total_unrealized_gain": 0,
                "total_unrealized_pct": 0,
                "total_realized_gain": 0,
                "total_short_term_gain": 0,
                "total_long_term_gain": 0,
            }
        }

    # Fetch live prices for all tickers
    tickers = [h.ticker for h in holdings]
    names = {h.ticker: h.name for h in holdings if h.name}
    try:
        current_prices = await market_data.get_current_prices(tickers, names=names)
    except Exception:
        current_prices = {}

    # Compute unrealized P&L per holding
    holding_pnl = []
    total_invested = 0.0
    total_current_value = 0.0

    for h in holdings:
        avg_cost = float(h.avg_cost)
        shares   = float(h.shares)
        price    = current_prices.get(h.ticker, 0.0)

        invested = avg_cost * shares
        total_invested += invested

        if price > 0:
            current_value  = price * shares
            unrealized     = current_value - invested
            pct            = (price - avg_cost) / avg_cost * 100 if avg_cost > 0 else 0.0
            total_current_value += current_value
        else:
            # LTP unavailable — assume break-even for portfolio summary totals
            current_value  = None
            unrealized     = None
            pct            = None
            total_current_value += invested  # treat as break-even so summary % isn't distorted

        holding_pnl.append({
            "id":              h.id,
            "ticker":          h.ticker,
            "name":            h.name,
            "shares":          shares,
            "avg_cost":        avg_cost,
            "current_price":   price if price > 0 else None,
            "invested":        round(invested, 2),
            "current_value":   round(current_value, 2) if current_value is not None else None,
            "unrealized_gain": round(unrealized, 2) if unrealized is not None else None,
            "unrealized_pct":  round(pct, 2) if pct is not None else None,
        })

    # Fetch realized P&L stored from broker uploads
    real_result = await db.execute(
        select(RealizedPnL).where(RealizedPnL.portfolio_id == portfolio_id)
    )
    realized_rows = real_result.scalars().all()
    total_stcg = sum(float(r.short_term_gain) for r in realized_rows)
    total_ltcg = sum(float(r.long_term_gain)  for r in realized_rows)
    total_realized = total_stcg + total_ltcg

    total_unrealized = total_current_value - total_invested
    total_unrealized_pct = (total_unrealized / total_invested * 100) if total_invested > 0 else 0.0

    return {
        "holdings": holding_pnl,
        "summary": {
            "total_invested":      round(total_invested, 2),
            "total_current_value": round(total_current_value, 2),
            "total_unrealized_gain": round(total_unrealized, 2),
            "total_unrealized_pct":  round(total_unrealized_pct, 2),
            "total_realized_gain":   round(total_realized, 2),
            "total_short_term_gain": round(total_stcg, 2),
            "total_long_term_gain":  round(total_ltcg, 2),
        },
    }


@router.post("/{portfolio_id}/holdings")
async def add_holding(
    portfolio_id: int, body: HoldingIn, db: AsyncSession = Depends(get_db)
):
    portfolio = await db.get(Portfolio, portfolio_id)
    if not portfolio:
        raise HTTPException(404, "Portfolio not found")

    holding = Holding(
        portfolio_id=portfolio_id,
        ticker=body.ticker.upper().strip(),
        shares=body.shares,
        avg_cost=body.avg_cost,
    )
    db.add(holding)
    await db.commit()
    await db.refresh(holding)
    return {
        "id": holding.id,
        "ticker": holding.ticker,
        "name": holding.name,
        "shares": float(holding.shares),
        "avg_cost": float(holding.avg_cost),
    }


@router.delete("/{portfolio_id}/holdings")
async def clear_all_holdings(portfolio_id: int, db: AsyncSession = Depends(get_db)):
    """Delete ALL holdings and realized P&L for a portfolio (fresh start)."""
    portfolio = await db.get(Portfolio, portfolio_id)
    if not portfolio:
        raise HTTPException(404, "Portfolio not found")

    h_result = await db.execute(select(Holding).where(Holding.portfolio_id == portfolio_id))
    holdings_count = 0
    for h in h_result.scalars().all():
        await db.delete(h)
        holdings_count += 1

    pnl_result = await db.execute(select(RealizedPnL).where(RealizedPnL.portfolio_id == portfolio_id))
    pnl_count = 0
    for r in pnl_result.scalars().all():
        await db.delete(r)
        pnl_count += 1

    await db.commit()
    return {"deleted_holdings": holdings_count, "deleted_pnl_rows": pnl_count}


@router.delete("/{portfolio_id}/holdings/{holding_id}")
async def delete_holding(
    portfolio_id: int, holding_id: int, db: AsyncSession = Depends(get_db)
):
    holding = await db.get(Holding, holding_id)
    if not holding or holding.portfolio_id != portfolio_id:
        raise HTTPException(404, "Holding not found")
    await db.delete(holding)
    await db.commit()
    return {"deleted": holding_id}


def _read_excel_rows(contents: bytes, filename: str) -> list[tuple]:
    """
    Read all rows from an Excel file, returning a list of value tuples.
    Supports both .xlsx (via openpyxl) and .xls (via xlrd).
    """
    fname = (filename or "").lower()
    if fname.endswith(".xls") and not fname.endswith(".xlsx"):
        import xlrd
        wb = xlrd.open_workbook(file_contents=contents)
        ws = wb.sheet_by_index(0)
        rows: list[tuple] = []
        for r in range(ws.nrows):
            row = []
            for c in range(ws.ncols):
                cell = ws.cell(r, c)
                if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                    row.append(None)
                elif cell.ctype == xlrd.XL_CELL_DATE:
                    row.append(xlrd.xldate_as_datetime(cell.value, wb.datemode))
                else:
                    row.append(cell.value)
            rows.append(tuple(row))
        return rows
    else:
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        return list(ws.iter_rows(values_only=True))


@router.post("/{portfolio_id}/holdings/upload")
async def upload_holdings_excel(
    portfolio_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """
    Supports two Excel formats:

    1. BROKER FORMAT (PROFITMART / similar broker reports):
       - Rows 1-3: metadata; Row 4: headers with Scrip_Symbol, Purchase_Qty, etc.
       - Automatically extracts open positions + realized P&L

    2. SIMPLE FORMAT:
       - Row 1: Ticker | Shares | Avg Cost
    """
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx and .xls files are supported")

    portfolio = await db.get(Portfolio, portfolio_id)
    if not portfolio:
        raise HTTPException(404, "Portfolio not found")

    contents = await file.read()
    try:
        rows = _read_excel_rows(contents, file.filename)
    except Exception:
        raise HTTPException(400, "Could not read Excel file. Make sure it is a valid .xlsx or .xls file.")

    if len(rows) < 2:
        raise HTTPException(400, "File must have a header row and at least one data row.")

    # Reject derivative files — use the dedicated /derivatives/upload endpoint
    if _is_derivative_format(rows):
        raise HTTPException(
            400,
            "This looks like a Derivatives P&L file. "
            "Please use the 'Derivatives' tab to upload it."
        )

    broker_fmt  = _is_broker_format(rows)
    client_info: dict = {}

    if broker_fmt:
        holdings_to_add, realized_to_add, skipped, client_info = _parse_broker_format(rows)
    else:
        # Simple format
        headers = _normalise_headers(rows[0])

        def find_col(candidates: list[str]) -> int:
            for c in candidates:
                if c in headers:
                    return headers.index(c)
            return -1

        ticker_idx = find_col(["ticker", "symbol", "stock", "scrip_symbol"])
        shares_idx = find_col(["shares", "quantity", "qty", "units", "purchase_qty"])
        cost_idx   = find_col(["avg_cost", "average_cost", "cost", "price",
                                "buy_price", "purchase_rate"])

        missing = []
        if ticker_idx == -1: missing.append("ticker/symbol")
        if shares_idx == -1: missing.append("shares/qty")
        if cost_idx   == -1: missing.append("avg_cost/price")
        if missing:
            raise HTTPException(
                400,
                f"Could not find required columns: {', '.join(missing)}. "
                f"Headers found: {', '.join(h for h in headers if h)}"
            )

        holdings_to_add, realized_to_add, skipped = [], [], []
        for i, row in enumerate(rows[1:], start=2):
            try:
                ticker   = str(row[ticker_idx]).upper().strip()
                shares   = float(row[shares_idx])
                avg_cost = float(row[cost_idx])
                if not ticker or shares <= 0 or avg_cost <= 0:
                    raise ValueError("Invalid values")
                holdings_to_add.append({"ticker": ticker, "shares": shares, "avg_cost": avg_cost})
            except Exception:
                skipped.append(f"row {i}")

    if not holdings_to_add and not realized_to_add:
        raise HTTPException(400, "No valid positions found in the file.")

    # ── Update client name + broker_client_id from the file ───────────────────
    broker_name = client_info.get("client_name", "").strip()
    broker_cid  = client_info.get("client_id",   "").strip()
    if broker_name:
        client_obj = await db.get(Client, portfolio.client_id)
        if client_obj:
            client_obj.name = broker_name
    if broker_cid:
        portfolio.broker_client_id = broker_cid
    # ─────────────────────────────────────────────────────────────────────────

    # ── Delete ALL existing holdings and realized P&L before re-importing ──────
    existing_holdings = await db.execute(
        select(Holding).where(Holding.portfolio_id == portfolio_id)
    )
    for h in existing_holdings.scalars().all():
        await db.delete(h)

    existing_pnl = await db.execute(
        select(RealizedPnL).where(RealizedPnL.portfolio_id == portfolio_id)
    )
    for row in existing_pnl.scalars().all():
        await db.delete(row)
    # ──────────────────────────────────────────────────────────────────────────

    # Insert open holdings
    added_tickers = []
    for h in holdings_to_add:
        db.add(Holding(
            portfolio_id=portfolio_id,
            ticker=h["ticker"],
            name=h.get("name"),
            shares=h["shares"],
            avg_cost=h["avg_cost"],
        ))
        added_tickers.append(h["ticker"])

    # Insert realized P&L
    if realized_to_add:
        for r in realized_to_add:
            db.add(RealizedPnL(
                portfolio_id=portfolio_id,
                ticker=r["ticker"],
                short_term_gain=r["short_term_gain"],
                long_term_gain=r["long_term_gain"],
            ))

    await db.commit()
    return {
        "added": len(added_tickers),
        "tickers": added_tickers,
        "realized_pnl_imported": len(realized_to_add),
        "skipped": skipped,
        "client_name": broker_name or None,
        "client_id":   broker_cid  or None,
        "format_detected": "broker" if broker_fmt else "simple",
    }


# ── Derivative Endpoints ────────────────────────────────────────────────────────

@router.post("/{portfolio_id}/derivatives/upload")
async def upload_derivatives_excel(
    portfolio_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """
    Upload a PROFITMART Derivative P&L Excel (DER P&L report).
    Replaces all existing derivative trades for this portfolio.
    """
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx and .xls files are supported")

    portfolio = await db.get(Portfolio, portfolio_id)
    if not portfolio:
        raise HTTPException(404, "Portfolio not found")

    contents = await file.read()
    try:
        rows = _read_excel_rows(contents, file.filename)
    except Exception:
        raise HTTPException(400, "Could not read Excel file.")

    if not _is_derivative_format(rows):
        raise HTTPException(
            400,
            "File does not appear to be a Derivative P&L report. "
            "Expected INSTRUMENT_TYPE and Booked P/L columns (PROFITMART DER P&L format)."
        )

    trades, skipped, client_info = _parse_derivative_format(rows)

    if not trades:
        raise HTTPException(400, "No valid derivative trades found in the file.")

    # ── Client ID cross-validation ─────────────────────────────────────────────
    deriv_cid  = client_info.get("client_id",   "").strip()
    deriv_name = client_info.get("client_name", "").strip()

    if deriv_cid and portfolio.broker_client_id:
        if str(deriv_cid) != str(portfolio.broker_client_id):
            raise HTTPException(
                400,
                f"Client ID mismatch: the uploaded portfolio file belongs to client "
                f"'{portfolio.broker_client_id}', but this Derivatives file belongs to "
                f"client '{deriv_cid}' ({deriv_name}). "
                f"Please upload the correct DER P&L file for the same client."
            )

    # ── Update client name from derivative file if not already set ────────────
    if deriv_name:
        client_obj = await db.get(Client, portfolio.client_id)
        if client_obj and client_obj.name in ("Demo Client", "Client", ""):
            client_obj.name = deriv_name
    if deriv_cid and not portfolio.broker_client_id:
        portfolio.broker_client_id = deriv_cid
    # ─────────────────────────────────────────────────────────────────────────

    # Delete all existing derivative trades for this portfolio
    existing = await db.execute(
        select(DerivativeTrade).where(DerivativeTrade.portfolio_id == portfolio_id)
    )
    for t in existing.scalars().all():
        await db.delete(t)

    # Insert new trades
    for t in trades:
        db.add(DerivativeTrade(portfolio_id=portfolio_id, **t))

    await db.commit()
    return {
        "imported": len(trades),
        "skipped":  skipped,
        "client_name": deriv_name or None,
        "client_id":   deriv_cid  or None,
    }


@router.get("/{portfolio_id}/derivatives/pnl")
async def get_derivatives_pnl(portfolio_id: int, db: AsyncSession = Depends(get_db)):
    """Returns aggregate F&O P&L stats for the portfolio."""
    portfolio = await db.get(Portfolio, portfolio_id)
    if not portfolio:
        raise HTTPException(404, "Portfolio not found")

    result = await db.execute(
        select(DerivativeTrade).where(DerivativeTrade.portfolio_id == portfolio_id)
    )
    trades = result.scalars().all()

    if not trades:
        return {
            "total_trades": 0,
            "total_booked_pnl": 0,
            "total_profit": 0,
            "total_loss": 0,
            "winning_trades": 0,
            "losing_trades": 0,
            "by_month": [],
            "by_option_type": [],
            "best_trade": None,
            "worst_trade": None,
        }

    total_pnl    = sum(float(t.booked_pnl)    for t in trades)
    total_profit = sum(float(t.booked_profit) for t in trades)
    total_loss   = sum(float(t.booked_loss)   for t in trades)
    winning      = sum(1 for t in trades if float(t.booked_pnl) > 0)
    losing       = sum(1 for t in trades if float(t.booked_pnl) < 0)

    # Monthly aggregation
    monthly: dict = defaultdict(lambda: {"pnl": 0.0, "profit": 0.0, "loss": 0.0, "trades": 0})
    for t in trades:
        key = t.trade_date.strftime("%Y-%m") if t.trade_date else "Unknown"
        monthly[key]["pnl"]    += float(t.booked_pnl)
        monthly[key]["profit"] += float(t.booked_profit)
        monthly[key]["loss"]   += float(t.booked_loss)
        monthly[key]["trades"] += 1

    by_month = [
        {
            "month":  k,
            "pnl":    round(v["pnl"], 2),
            "profit": round(v["profit"], 2),
            "loss":   round(v["loss"], 2),
            "trades": v["trades"],
        }
        for k, v in sorted(monthly.items())
    ]

    # CE / PE aggregation
    by_type_agg: dict = defaultdict(lambda: {"pnl": 0.0, "trades": 0})
    for t in trades:
        ot = t.option_type or "Other"
        by_type_agg[ot]["pnl"]    += float(t.booked_pnl)
        by_type_agg[ot]["trades"] += 1

    by_option_type = [
        {"type": k, "pnl": round(v["pnl"], 2), "trades": v["trades"]}
        for k, v in sorted(by_type_agg.items())
    ]

    # Best / worst single trades
    sorted_by_pnl = sorted(trades, key=lambda t: float(t.booked_pnl))
    worst = sorted_by_pnl[0]
    best  = sorted_by_pnl[-1]

    def trade_dict(t) -> dict:
        return {
            "scrip_symbol": t.scrip_symbol,
            "option_type":  t.option_type,
            "underlying":   t.underlying,
            "strike_price": float(t.strike_price) if t.strike_price else None,
            "trade_date":   t.trade_date.strftime("%d-%b-%Y") if t.trade_date else None,
            "buy_qty":      float(t.buy_qty),
            "buy_rate":     float(t.buy_rate),
            "sell_rate":    float(t.sell_rate),
            "booked_pnl":   round(float(t.booked_pnl), 2),
        }

    return {
        "total_trades":     len(trades),
        "total_booked_pnl": round(total_pnl, 2),
        "total_profit":     round(total_profit, 2),
        "total_loss":       round(total_loss, 2),
        "winning_trades":   winning,
        "losing_trades":    losing,
        "by_month":         by_month,
        "by_option_type":   by_option_type,
        "best_trade":       trade_dict(best),
        "worst_trade":      trade_dict(worst),
    }


@router.delete("/{portfolio_id}/derivatives")
async def clear_derivatives(portfolio_id: int, db: AsyncSession = Depends(get_db)):
    """Delete all derivative trades for a portfolio."""
    portfolio = await db.get(Portfolio, portfolio_id)
    if not portfolio:
        raise HTTPException(404, "Portfolio not found")

    result = await db.execute(
        select(DerivativeTrade).where(DerivativeTrade.portfolio_id == portfolio_id)
    )
    count = 0
    for t in result.scalars().all():
        await db.delete(t)
        count += 1
    await db.commit()
    return {"deleted": count}


# ── Tax Report ─────────────────────────────────────────────────────────────────

def _get_fy_dividends(holdings) -> list[dict]:
    """
    Fetch FY 2025-26 dividend events for each open holding using yfinance.
    Returns rows sorted by (ticker, ex_date):
        [{ticker, display, ticker_clean, ex_date, dps, shares, total}, ...]

    NOTE: Uses current open-position share count as proxy for shares held on
    ex-dividend date (individual buy/sell dates are not stored per-transaction).
    """
    import yfinance as yf
    import pandas as pd

    fy_start = pd.Timestamp("2025-04-01")
    fy_end   = pd.Timestamp("2026-03-31")
    rows: list[dict] = []

    for h in holdings:
        try:
            divs = yf.Ticker(h.ticker).dividends
            if divs is None or divs.empty:
                continue
            # Strip timezone so comparison with tz-naive Timestamps works
            if divs.index.tz is not None:
                divs.index = divs.index.tz_convert("UTC").tz_localize(None)
            fy_divs = divs[(divs.index >= fy_start) & (divs.index <= fy_end)]
            if fy_divs.empty:
                continue
            shares       = float(h.shares)
            display      = h.name or h.ticker.replace(".NS", "").replace(".BO", "")
            ticker_clean = h.ticker.replace(".NS", "").replace(".BO", "")
            for ex_dt, dps in fy_divs.items():
                dps_val = round(float(dps), 4)
                rows.append({
                    "ticker":       h.ticker,
                    "display":      display,
                    "ticker_clean": ticker_clean,
                    "ex_date":      ex_dt.date(),
                    "dps":          dps_val,
                    "shares":       shares,
                    "total":        round(dps_val * shares, 2),
                })
        except Exception:
            pass

    return sorted(rows, key=lambda r: (r["ticker"], str(r["ex_date"])))


def _tax_excel(
    client_name: str,
    holdings,
    realized_rows,
    prices: dict,
    deriv_summary: dict | None = None,
    dividend_rows: list | None = None,
) -> openpyxl.Workbook:
    """
    Generate a tax filing workbook with sheets:
      1. Tax Summary        – AY, STCG/LTCG totals, F&O, dividend, estimated tax
      2. Realized Gains     – Per-stock STCG and LTCG breakdown
      3. Open Positions     – Unrealized P&L on current holdings
      4. F&O Trades         – Monthly derivative P&L (if deriv_summary provided)
      5. Dividend Income    – Per-stock, per-ex-date dividend detail (if dividend_rows provided)
    """
    wb = openpyxl.Workbook()

    def _add_logo(target_ws) -> None:
        """Stamp the PROFITMART logo at B1 of any sheet (each call needs a fresh Image object)."""
        try:
            logo_img = XLImage(io.BytesIO(_PROFITMART_LOGO_BYTES))
            logo_img.width  = 150
            logo_img.height = 30
            logo_img.anchor = "B1"
            target_ws.add_image(logo_img)
            target_ws.row_dimensions[1].height = 38
        except Exception:
            pass

    # ── Colour / font helpers ──────────────────────────────────────────────────
    NAVY   = "1B3A6B"
    BLUE   = "2E6DB4"
    LBLUE  = "D9E8F7"
    GREEN  = "006B3C"
    RED    = "C00000"
    GRAY   = "F5F5F5"
    WHITE  = "FFFFFF"
    BLACK  = "000000"
    ORANGE = "7B3800"

    def hdr_font(bold=True, color=WHITE, sz=11):
        return Font(bold=bold, color=color, size=sz, name="Calibri")

    def cell_font(bold=False, color=BLACK, sz=10):
        return Font(bold=bold, color=color, size=sz, name="Calibri")

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def thin_border():
        s = Side(style="thin", color="C0C0C0")
        return Border(left=s, right=s, top=s, bottom=s)

    def center(wrap=False):
        return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

    def right_align():
        return Alignment(horizontal="right", vertical="center")

    def inr(ws, row, col, value, color=None):
        cell = ws.cell(row=row, column=col, value=round(float(value), 2))
        cell.number_format = u'₹#,##0.00'
        cell.alignment = right_align()
        cell.border = thin_border()
        cell.font = cell_font(color=color or BLACK)
        return cell

    def label(ws, row, col, text, bold=False, bg=None, fg=BLACK, wrap=False, colspan=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = cell_font(bold=bold, color=fg, sz=10)
        cell.alignment = center(wrap) if colspan > 1 else Alignment(vertical="center", wrap_text=wrap)
        cell.border = thin_border()
        if bg:
            cell.fill = fill(bg)
        if colspan > 1:
            ws.merge_cells(start_row=row, start_column=col,
                           end_row=row, end_column=col + colspan - 1)
        return cell

    FY = "2025-26"
    AY = "2026-27"
    today = date_type.today().strftime("%d-%b-%Y")

    # ── Compute equity totals ──────────────────────────────────────────────────
    total_stcg = sum(float(r.short_term_gain) for r in realized_rows)
    total_ltcg = sum(float(r.long_term_gain)  for r in realized_rows)
    ltcg_exempt   = min(max(total_ltcg, 0), 125_000)
    ltcg_taxable  = max(total_ltcg - ltcg_exempt, 0)
    tax_stcg      = max(total_stcg, 0) * 0.20
    tax_ltcg      = ltcg_taxable * 0.125
    total_tax_est = tax_stcg + tax_ltcg

    total_invested = sum(float(h.avg_cost) * float(h.shares) for h in holdings)
    priced = [(h, prices.get(h.ticker, 0.0)) for h in holdings]
    total_current  = sum(p * float(h.shares) for h, p in priced if p > 0)
    total_unrealized = sum(
        (p - float(h.avg_cost)) * float(h.shares)
        for h, p in priced if p > 0
    )

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 – Tax Summary
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Tax Summary"
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 22
    ws.row_dimensions[1].height = 8

    _add_logo(ws)   # Sheet 1 logo

    # Title banner
    ws.row_dimensions[2].height = 30
    ws.merge_cells("B2:C2")
    t = ws["B2"]
    t.value = f"Capital Gains & F&O Tax Report — FY {FY}  |  AY {AY}"
    t.font  = Font(bold=True, color=WHITE, size=14, name="Calibri")
    t.fill  = fill(NAVY)
    t.alignment = center()
    t.border = thin_border()

    # Meta info
    for r, (lbl, val) in enumerate([
        ("Client Name",       client_name),
        ("Assessment Year",   f"AY {AY}"),
        ("Financial Year",    f"FY {FY}  (01-Apr-2025 to 31-Mar-2026)"),
        ("Report Generated",  today),
    ], start=3):
        ws.row_dimensions[r].height = 18
        label(ws, r, 2, lbl, bold=True, bg=LBLUE)
        label(ws, r, 3, val)

    # Spacer row 7
    ws.row_dimensions[7].height = 10

    # ── Section A: Realized Capital Gains ─────────────────────────────────────
    ws.row_dimensions[8].height = 20
    ws.merge_cells("B8:C8")
    sh = ws["B8"]
    sh.value = "Section A — Realized Capital Gains (Equity)"
    sh.font, sh.fill, sh.alignment, sh.border = hdr_font(), fill(BLUE), center(), thin_border()

    rows_a = [
        ("Short-Term Capital Gain / Loss (STCG)",    total_stcg),
        ("Long-Term Capital Gain / Loss (LTCG)",     total_ltcg),
        ("LTCG Exemption (up to ₹1,25,000)",        -ltcg_exempt),
        ("Net Taxable LTCG",                         ltcg_taxable),
    ]
    for i, (lbl_text, val) in enumerate(rows_a, start=9):
        ws.row_dimensions[i].height = 18
        label(ws, i, 2, lbl_text, bg=GRAY)
        color = GREEN if val >= 0 else RED
        inr(ws, i, 3, val, color=color)

    # ── Section B: Tax Estimate ────────────────────────────────────────────────
    ws.row_dimensions[13].height = 10
    ws.row_dimensions[14].height = 20
    ws.merge_cells("B14:C14")
    sh2 = ws["B14"]
    sh2.value = "Section B — Estimated Tax Liability (Equity Capital Gains)"
    sh2.font, sh2.fill, sh2.alignment, sh2.border = hdr_font(), fill(BLUE), center(), thin_border()

    tax_rows = [
        ("Tax on STCG @ 20%  (Section 111A)",   tax_stcg),
        ("Tax on LTCG @ 12.5%  (Section 112A)", tax_ltcg),
        ("Total Estimated Tax (Equity)",         total_tax_est),
    ]
    for i, (lbl_text, val) in enumerate(tax_rows, start=15):
        ws.row_dimensions[i].height = 18
        bold = (i == 17)
        label(ws, i, 2, lbl_text, bold=bold, bg=GRAY if not bold else LBLUE)
        inr(ws, i, 3, val, color=RED if val > 0 else GREEN)
        if bold:
            ws.cell(row=i, column=2).font = cell_font(bold=True)

    # ── Section C: Open Positions Summary ─────────────────────────────────────
    ws.row_dimensions[18].height = 10
    ws.row_dimensions[19].height = 20
    ws.merge_cells("B19:C19")
    sh3 = ws["B19"]
    sh3.value = "Section C — Open Positions (Unrealized)"
    sh3.font, sh3.fill, sh3.alignment, sh3.border = hdr_font(), fill(BLUE), center(), thin_border()

    open_rows = [
        ("Total Amount Invested (Open Positions)",   total_invested),
        ("Current Portfolio Value (Live Prices)",    total_current),
        ("Unrealized Gain / Loss",                   total_unrealized),
    ]
    for i, (lbl_text, val) in enumerate(open_rows, start=20):
        ws.row_dimensions[i].height = 18
        label(ws, i, 2, lbl_text, bg=GRAY)
        inr(ws, i, 3, val, color=GREEN if val >= 0 else RED)

    next_row = 23  # tracks next available row

    # ── Section D: F&O P&L (optional) ─────────────────────────────────────────
    if deriv_summary:
        d_net    = deriv_summary["total_pnl"]
        d_profit = deriv_summary["total_profit"]
        d_loss   = deriv_summary["total_loss"]
        d_trades = deriv_summary["total_trades"]

        ws.row_dimensions[next_row].height = 10  # spacer
        next_row += 1

        ws.row_dimensions[next_row].height = 20
        ws.merge_cells(f"B{next_row}:C{next_row}")
        sh4 = ws[f"B{next_row}"]
        sh4.value = "Section D — F&O / Derivatives Trading P&L"
        sh4.font, sh4.fill, sh4.alignment, sh4.border = hdr_font(), fill(BLUE), center(), thin_border()
        next_row += 1

        fo_rows = [
            ("Gross Booked Profit (F&O)",  d_profit),
            ("Gross Booked Loss (F&O)",   -d_loss),
            ("Net F&O P&L",               d_net),
        ]
        for lbl_text, val in fo_rows:
            ws.row_dimensions[next_row].height = 18
            label(ws, next_row, 2, lbl_text, bg=GRAY)
            inr(ws, next_row, 3, val, color=GREEN if val >= 0 else RED)
            next_row += 1

        # Informational note
        ws.row_dimensions[next_row].height = 20
        ws.merge_cells(f"B{next_row}:C{next_row}")
        note4 = ws[f"B{next_row}"]
        note4.value = (
            f"ℹ  F&O income ({d_trades} trades) is taxable as Business Income "
            "(non-speculative) under head PGBP. Consult a CA for ITR-3 filing."
        )
        note4.font = Font(italic=True, color=ORANGE, size=9, name="Calibri")
        note4.alignment = Alignment(wrap_text=True, vertical="center")
        next_row += 1

    # ── Section E: Dividend Income ─────────────────────────────────────────────
    if dividend_rows:
        total_div_e = sum(r["total"] for r in dividend_rows)

        ws.row_dimensions[next_row].height = 10  # spacer
        next_row += 1

        ws.row_dimensions[next_row].height = 20
        ws.merge_cells(f"B{next_row}:C{next_row}")
        sh_e = ws[f"B{next_row}"]
        sh_e.value = "Section E — Dividend Income (FY 2025-26)"
        sh_e.font, sh_e.fill, sh_e.alignment, sh_e.border = (
            hdr_font(), fill(BLUE), center(), thin_border()
        )
        next_row += 1

        div_summary_rows_e = [
            ("Total Dividend Income Received",           total_div_e),
            ("Indicative TDS @ 10% (if >₹5,000/co.)",   round(total_div_e * 0.10, 2)),
        ]
        for lbl_text, val in div_summary_rows_e:
            ws.row_dimensions[next_row].height = 18
            label(ws, next_row, 2, lbl_text, bg=GRAY)
            inr(ws, next_row, 3, val, color=GREEN if val >= 0 else RED)
            next_row += 1

        ws.row_dimensions[next_row].height = 24
        ws.merge_cells(f"B{next_row}:C{next_row}")
        note_e = ws[f"B{next_row}"]
        note_e.value = (
            "ℹ  Post Finance Act 2020, dividends are fully taxable at slab rates. "
            "TDS @ 10% is deducted when total dividend from one company exceeds ₹5,000 in a FY. "
            "See Sheet 5 for stock-wise detail."
        )
        note_e.font = Font(italic=True, color=ORANGE, size=9, name="Calibri")
        note_e.alignment = Alignment(wrap_text=True, vertical="center")
        next_row += 1

    # Disclaimer
    ws.row_dimensions[next_row].height = 10    # spacer
    next_row += 1
    ws.row_dimensions[next_row].height = 30
    ws.merge_cells(f"B{next_row}:C{next_row}")
    disc = ws[f"B{next_row}"]
    disc.value = (
        "⚠  Disclaimer: Tax rates as per Finance Act 2024 (STCG 20%, LTCG 12.5% with ₹1.25L "
        "exemption). Figures are computed from broker-uploaded data. Please verify with a "
        "Chartered Accountant before filing your ITR."
    )
    disc.font      = Font(italic=True, color="7F7F7F", size=9, name="Calibri")
    disc.alignment = Alignment(wrap_text=True, vertical="center")

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 – Realized Capital Gains (per stock)
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Realized Capital Gains")
    COL_WIDTHS2 = [2, 32, 16, 18, 18, 20, 20]
    for i, w in enumerate(COL_WIDTHS2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    _add_logo(ws2)  # Sheet 2 logo

    ws2.row_dimensions[2].height = 26
    ws2.merge_cells("B2:G2")
    t2 = ws2["B2"]
    t2.value = f"Realized Capital Gains — FY {FY}  (Equity)"
    t2.font, t2.fill, t2.alignment, t2.border = hdr_font(sz=12), fill(NAVY), center(), thin_border()

    headers2 = ["#", "Stock Name", "Ticker", "STCG (₹)", "LTCG (₹)", "Tax Est. STCG", "Tax Est. LTCG"]
    ws2.row_dimensions[3].height = 20
    for col, hdr in enumerate(headers2, 2):
        c = ws2.cell(row=3, column=col, value=hdr)
        c.font, c.fill, c.alignment, c.border = hdr_font(), fill(BLUE), center(True), thin_border()

    row = 4
    for idx, r in enumerate(realized_rows, 1):
        stcg = float(r.short_term_gain)
        ltcg = float(r.long_term_gain)
        t_stcg = max(stcg, 0) * 0.20
        t_ltcg = max(ltcg, 0) * 0.125
        bg = WHITE if idx % 2 == 0 else GRAY

        h_match = next((h for h in holdings if h.ticker == r.ticker), None)
        dname = (h_match.name or r.ticker) if h_match else r.ticker
        ticker_display = r.ticker.replace(".NS", "").replace(".BO", "")

        ws2.row_dimensions[row].height = 18
        for col, val in enumerate([idx, dname, ticker_display, stcg, ltcg, t_stcg, t_ltcg], 2):
            c = ws2.cell(row=row, column=col, value=val)
            c.border = thin_border()
            c.fill   = fill(bg)
            if col in (5, 6, 7, 8):
                c.number_format = u'₹#,##0.00'
                c.alignment = right_align()
                c.font = cell_font(color=GREEN if float(val) >= 0 else RED)
            else:
                c.font = cell_font()
        row += 1

    # Totals row
    ws2.row_dimensions[row].height = 20
    total_labels = ["", "TOTAL", "", total_stcg, total_ltcg,
                    max(total_stcg, 0) * 0.20, max(total_ltcg - 125_000, 0) * 0.125]
    for col, val in enumerate(total_labels, 2):
        c = ws2.cell(row=row, column=col, value=val)
        c.font, c.fill, c.border = hdr_font(color=WHITE), fill(NAVY), thin_border()
        if col >= 5:
            c.number_format = u'₹#,##0.00'
            c.alignment = right_align()

    row += 2
    ws2.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    note = ws2.cell(row=row, column=2,
                    value="Tax Rates: STCG @ 20% (Sec 111A) | LTCG @ 12.5% on gains above ₹1,25,000 (Sec 112A)  |  Finance Act 2024")
    note.font = Font(italic=True, size=9, color="595959", name="Calibri")

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3 – Open Positions
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Open Positions")
    COL_WIDTHS3 = [2, 32, 14, 10, 16, 16, 18, 18, 14]
    for i, w in enumerate(COL_WIDTHS3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    _add_logo(ws3)  # Sheet 3 logo

    ws3.row_dimensions[2].height = 26
    ws3.merge_cells("B2:J2")
    t3 = ws3["B2"]
    t3.value = f"Open Positions as of {today}  |  FY {FY}"
    t3.font, t3.fill, t3.alignment, t3.border = hdr_font(sz=12), fill(NAVY), center(), thin_border()

    headers3 = ["#", "Stock Name", "Ticker", "Qty", "Avg Cost (₹)", "LTP (₹)",
                "Invested (₹)", "Current Value (₹)", "Unrealized P&L (₹)", "P&L %"]
    ws3.row_dimensions[3].height = 20
    for col, hdr in enumerate(headers3, 2):
        c = ws3.cell(row=3, column=col, value=hdr)
        c.font, c.fill, c.alignment, c.border = hdr_font(), fill(BLUE), center(True), thin_border()

    row3 = 4
    for idx, h in enumerate(holdings, 1):
        ltp      = prices.get(h.ticker, 0.0)
        invested = float(h.avg_cost) * float(h.shares)
        cur_val  = ltp * float(h.shares) if ltp > 0 else None
        unreal   = (cur_val - invested) if cur_val is not None else None
        pct      = ((ltp - float(h.avg_cost)) / float(h.avg_cost) * 100) if ltp > 0 and float(h.avg_cost) > 0 else None
        bg       = WHITE if idx % 2 == 0 else GRAY
        name_disp   = h.name or h.ticker
        ticker_disp = h.ticker.replace(".NS", "").replace(".BO", "")

        ws3.row_dimensions[row3].height = 18
        vals = [idx, name_disp, ticker_disp, float(h.shares), float(h.avg_cost),
                ltp if ltp > 0 else "—", invested, cur_val or "—", unreal or "—",
                f"{pct:+.2f}%" if pct is not None else "—"]
        num_cols = {5, 6, 7, 8, 9}
        for col, val in enumerate(vals, 2):
            c = ws3.cell(row=row3, column=col, value=val)
            c.border = thin_border()
            c.fill   = fill(bg)
            if col in num_cols and isinstance(val, (int, float)):
                c.number_format = u'₹#,##0.00'
                c.alignment = right_align()
                if col in (9, 10) and isinstance(val, float):
                    c.font = cell_font(color=GREEN if val >= 0 else RED)
                else:
                    c.font = cell_font()
            elif col == 4:
                c.alignment = right_align()
                c.font = cell_font()
            else:
                c.font = cell_font()
        row3 += 1

    ws3.row_dimensions[row3].height = 20
    tot_vals = ["", "TOTAL", "", "", "", "", total_invested, total_current or "—",
                total_unrealized or "—", ""]
    for col, val in enumerate(tot_vals, 2):
        c = ws3.cell(row=row3, column=col, value=val)
        c.font, c.fill, c.border = hdr_font(color=WHITE), fill(NAVY), thin_border()
        if col in (8, 9, 10) and isinstance(val, (int, float)):
            c.number_format = u'₹#,##0.00'
            c.alignment = right_align()

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 4 – F&O Trades (if derivative data available)
    # ══════════════════════════════════════════════════════════════════════════
    if deriv_summary and deriv_summary.get("by_month"):
        ws4 = wb.create_sheet("F&O Summary")
        ws4.column_dimensions["A"].width = 2
        ws4.column_dimensions["B"].width = 14
        ws4.column_dimensions["C"].width = 18
        ws4.column_dimensions["D"].width = 18
        ws4.column_dimensions["E"].width = 18
        ws4.column_dimensions["F"].width = 10
        _add_logo(ws4)  # Sheet 4 logo

        ws4.row_dimensions[2].height = 26
        ws4.merge_cells("B2:F2")
        t4 = ws4["B2"]
        t4.value = f"F&O Monthly P&L Summary — FY {FY}"
        t4.font, t4.fill, t4.alignment, t4.border = hdr_font(sz=12), fill(NAVY), center(), thin_border()

        ws4.row_dimensions[3].height = 20
        for col, hdr in enumerate(["Month", "Gross Profit (₹)", "Gross Loss (₹)", "Net P&L (₹)", "Trades"], 2):
            c = ws4.cell(row=3, column=col, value=hdr)
            c.font, c.fill, c.alignment, c.border = hdr_font(), fill(BLUE), center(True), thin_border()

        row4 = 4
        for idx, m in enumerate(deriv_summary["by_month"], 1):
            bg = WHITE if idx % 2 == 0 else GRAY
            ws4.row_dimensions[row4].height = 18
            net_pnl = m["pnl"]
            for col, val in enumerate([m["month"], m["profit"], m["loss"], net_pnl, m["trades"]], 2):
                c = ws4.cell(row=row4, column=col, value=val)
                c.border = thin_border()
                c.fill   = fill(bg)
                if col in (3, 4, 5) and isinstance(val, (int, float)):
                    c.number_format = u'₹#,##0.00'
                    c.alignment = right_align()
                    c.font = cell_font(color=(GREEN if val >= 0 else RED) if col == 5 else BLACK)
                else:
                    c.font = cell_font()
            row4 += 1

        # Totals
        ws4.row_dimensions[row4].height = 20
        tot4 = ["TOTAL", deriv_summary["total_profit"], deriv_summary["total_loss"],
                deriv_summary["total_pnl"], deriv_summary["total_trades"]]
        for col, val in enumerate(tot4, 2):
            c = ws4.cell(row=row4, column=col, value=val)
            c.font, c.fill, c.border = hdr_font(color=WHITE), fill(NAVY), thin_border()
            if col in (3, 4, 5) and isinstance(val, (int, float)):
                c.number_format = u'₹#,##0.00'
                c.alignment = right_align()

        row4 += 2
        ws4.merge_cells(start_row=row4, start_column=2, end_row=row4, end_column=6)
        note4 = ws4.cell(row=row4, column=2,
                         value="F&O income is taxable under Business Income (PGBP) — non-speculative. File ITR-3. Consult CA.")
        note4.font = Font(italic=True, size=9, color="595959", name="Calibri")

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 5 – Dividend Income (per stock, per ex-date)
    # ══════════════════════════════════════════════════════════════════════════
    if dividend_rows:
        ws5 = wb.create_sheet("Dividend Income")
        # Col widths: A(spacer), B(#), C(Stock Name), D(Ticker), E(Ex-Date),
        #             F(Shares), G(Div/Share), H(Total Div)
        COL_WIDTHS5 = [2, 4, 32, 14, 14, 12, 16, 18]
        for i, w in enumerate(COL_WIDTHS5, 1):
            ws5.column_dimensions[get_column_letter(i)].width = w
        _add_logo(ws5)  # Sheet 5 logo

        ws5.row_dimensions[2].height = 26
        ws5.merge_cells("B2:H2")
        t5 = ws5["B2"]
        t5.value = f"Dividend Income — FY {FY}"
        t5.font, t5.fill, t5.alignment, t5.border = (
            hdr_font(sz=12), fill(NAVY), center(), thin_border()
        )

        headers5 = ["#", "Stock Name", "Ticker", "Ex-Date",
                    "Shares Held", "Div/Share (₹)", "Total Dividend (₹)"]
        ws5.row_dimensions[3].height = 20
        for col, hdr in enumerate(headers5, 2):
            c = ws5.cell(row=3, column=col, value=hdr)
            c.font, c.fill, c.alignment, c.border = (
                hdr_font(), fill(BLUE), center(True), thin_border()
            )

        row5 = 4
        for idx, dr in enumerate(dividend_rows, 1):
            bg = WHITE if idx % 2 == 0 else GRAY
            ws5.row_dimensions[row5].height = 18
            vals5 = [
                idx,
                dr["display"],
                dr["ticker_clean"],
                str(dr["ex_date"]),
                dr["shares"],
                dr["dps"],
                dr["total"],
            ]
            for col, val in enumerate(vals5, 2):
                c = ws5.cell(row=row5, column=col, value=val)
                c.border = thin_border()
                c.fill   = fill(bg)
                if col in (7, 8) and isinstance(val, (int, float)):
                    c.number_format = u'₹#,##0.00'
                    c.alignment = right_align()
                    c.font = cell_font(color=GREEN)
                elif col == 6 and isinstance(val, (int, float)):
                    c.number_format = u'₹#,##0.00'
                    c.alignment = right_align()
                    c.font = cell_font()
                elif col == 5 and isinstance(val, (int, float)):
                    c.alignment = right_align()
                    c.font = cell_font()
                else:
                    c.font = cell_font()
            row5 += 1

        # Totals row
        total_div5 = sum(r["total"] for r in dividend_rows)
        ws5.row_dimensions[row5].height = 20
        tot5 = ["", "TOTAL", "", "", "", "", total_div5]
        for col, val in enumerate(tot5, 2):
            c = ws5.cell(row=row5, column=col, value=val)
            c.font, c.fill, c.border = hdr_font(color=WHITE), fill(NAVY), thin_border()
            if col == 8 and isinstance(val, (int, float)):
                c.number_format = u'₹#,##0.00'
                c.alignment = right_align()

        row5 += 2
        ws5.merge_cells(start_row=row5, start_column=2, end_row=row5, end_column=8)
        note5 = ws5.cell(row=row5, column=2, value=(
            "Note: Share count reflects current open positions as a proxy for shares held on "
            "ex-dividend date. Verify actual holding with broker contract notes. "
            "Dividend income is taxable at slab rate (post Finance Act 2020). "
            "Dividend data sourced via yfinance / exchange feeds — verify with company announcements."
        ))
        note5.font = Font(italic=True, size=9, color="595959", name="Calibri")
        note5.alignment = Alignment(wrap_text=True, vertical="center")
        ws5.row_dimensions[row5].height = 42

    return wb


@router.get("/{portfolio_id}/tax-report")
async def download_tax_report(
    portfolio_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Generate and download a Tax P&L report in Excel format for the current FY."""
    portfolio = await db.get(Portfolio, portfolio_id)
    if not portfolio:
        raise HTTPException(404, "Portfolio not found")

    client = await db.get(Client, portfolio.client_id)
    client_name = client.name if client else "Client"

    # Holdings
    h_res = await db.execute(select(Holding).where(Holding.portfolio_id == portfolio_id))
    holdings = h_res.scalars().all()

    # Realized P&L rows
    pnl_res = await db.execute(select(RealizedPnL).where(RealizedPnL.portfolio_id == portfolio_id))
    realized_rows = pnl_res.scalars().all()

    # Live prices for open positions
    tickers = [h.ticker for h in holdings]
    names   = {h.ticker: h.name for h in holdings if h.name}
    try:
        prices = await market_data.get_current_prices(tickers, names=names)
    except Exception:
        prices = {}

    # Derivative summary for F&O sheet
    deriv_res = await db.execute(
        select(DerivativeTrade).where(DerivativeTrade.portfolio_id == portfolio_id)
    )
    deriv_trades = deriv_res.scalars().all()
    deriv_summary = None
    if deriv_trades:
        monthly: dict = defaultdict(lambda: {"pnl": 0.0, "profit": 0.0, "loss": 0.0, "trades": 0})
        for t in deriv_trades:
            key = t.trade_date.strftime("%Y-%m") if t.trade_date else "Unknown"
            monthly[key]["pnl"]    += float(t.booked_pnl)
            monthly[key]["profit"] += float(t.booked_profit)
            monthly[key]["loss"]   += float(t.booked_loss)
            monthly[key]["trades"] += 1
        deriv_summary = {
            "total_pnl":    round(sum(float(t.booked_pnl)    for t in deriv_trades), 2),
            "total_profit": round(sum(float(t.booked_profit) for t in deriv_trades), 2),
            "total_loss":   round(sum(float(t.booked_loss)   for t in deriv_trades), 2),
            "total_trades": len(deriv_trades),
            "by_month": [
                {"month": k, "pnl": round(v["pnl"], 2), "profit": round(v["profit"], 2),
                 "loss": round(v["loss"], 2), "trades": v["trades"]}
                for k, v in sorted(monthly.items())
            ],
        }


    # Fetch FY 2025-26 dividend data (synchronous yfinance — run in thread pool)
    try:
        dividend_rows = await asyncio.to_thread(_get_fy_dividends, holdings)
    except Exception:
        dividend_rows = []

    wb = _tax_excel(
        client_name, holdings, realized_rows, prices,
        deriv_summary=deriv_summary,
        dividend_rows=dividend_rows or None,
    )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = re.sub(r"[^\w\-]", "_", portfolio.name)
    filename  = f"TaxReport_FY2025-26_{safe_name}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
